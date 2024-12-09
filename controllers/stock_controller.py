from flask import Blueprint, request, jsonify, send_from_directory,send_file
import openpyxl
from werkzeug.utils import secure_filename
import os
import math
import locale
from db import get_db_connection
from psycopg2.extras import RealDictCursor
request, jsonify, send_from_directory
from io import BytesIO
from openpyxl import Workbook



locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
UPLOAD_FOLDER=os.getenv("UPLOAD_FOLDER")
ALLOWED_EXTENSIONS = {'jpeg', 'jpg', 'png', 'gif'}
URL=os.getenv("URL")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



def upload_image():
    if 'image' not in request.files:
        return jsonify({"message": "No file part in the request"}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        filepath=URL+filename
        normalized_path = filepath.replace("\\", "/")
        return jsonify({"filepath": normalized_path}), 200

    return jsonify({"message": "File type not allowed. Only images are permitted"}), 400


def get_image(filename):
    try:
        print
        return send_from_directory(UPLOAD_FOLDER, filename)
    except FileNotFoundError:
        return jsonify({"message": "Image not found"}), 404




def find_products_by_product_kind():
    conn = get_db_connection() 
    try:
       
        id = request.args.get('id')
        productname = request.args.get('productname')
        sku = request.args.get('sku')
        colors = request.args.get('colors')
        size = request.args.get('size')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))

        
        base_query = """
            SELECT pk.*,
                   json_agg(p.*) AS product_list
            FROM productkinds pk
            LEFT JOIN product_list p ON pk.id = p.product_t_id
        """

       
        where_clauses = []
        query_params = []

        if id:
            where_clauses.append("pk.id = %s")
            query_params.append(id)
        if productname:
            where_clauses.append("pk.productname ILIKE %s")
            query_params.append(f"%{productname}%")
        if sku:
            where_clauses.append("pk.sku ILIKE %s")
            query_params.append(f"%{sku}%")
        if colors:
            where_clauses.append("%s = ANY(pk.colors)")
            query_params.append(colors)
        if size:
            where_clauses.append("%s = ANY(pk.size)")
            query_params.append(size)

        where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

      
        offset = (page - 1) * limit
        pagination_clause = f"LIMIT %s OFFSET %s"
        query_params.extend([limit, offset])

        
        final_query = f"""
            {base_query}
            {where_clause}
            GROUP BY pk.id
            ORDER BY pk.id ASC
            {pagination_clause}
        """

        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            
            cursor.execute(final_query, query_params)
            product_kinds = cursor.fetchall()

            
            count_query = f"SELECT COUNT(*) AS total FROM productkinds pk {where_clause}"
            cursor.execute(count_query, query_params[:-2]) 
            total_items = cursor.fetchone()["total"]

        
        formatted_data = [
            {
                "productKind": {
                    "id": pk["id"],
                    "productname": pk["productname"],
                    "sku": pk["sku"],
                    "colors": pk["colors"],
                    "size": pk["size"],
                    "description": pk["description"],
                },
                "productArray": pk["product_list"] or []
            }
            for pk in product_kinds
        ]

        response = {
            "data": formatted_data,
            "currentPage": page,
            "totalPages": math.ceil(total_items / limit),
        }

        return jsonify(response), 200

    except Exception as e:
        print(f"Error fetching product types and related products: {e}")
        return jsonify({"message": "Some error occurred!", "error": str(e)}), 500
    finally:
        conn.close()


def create_or_update_product_with_variants(id):
    conn = get_db_connection()  
    try:
        data = request.get_json()
        product_data = data.get('productData')
        generated_variants = data.get('generatedVariants', [])
        delete_variant_ids = data.get('deleteVariantIds', [])
        
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            conn.autocommit = False  
            
            main_product_id = None
            
           
            if id:
                update_product_query = """
                    UPDATE productkinds
                    SET productname = %s, description = %s, colors = %s, size = %s, sku = %s
                    WHERE id = %s
                    RETURNING id;
                """
                cursor.execute(update_product_query, [
                    product_data.get('productname'),
                    product_data.get('description'),
                    product_data.get('colors'),
                    product_data.get('size'),
                    product_data.get('sku'),
                    id
                ])
                result = cursor.fetchone()
                main_product_id = result["id"] if result else None
            
        
            if not main_product_id:
                insert_product_query = """
                    INSERT INTO productkinds (productname, description, colors, size, sku)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id;
                """
                cursor.execute(insert_product_query, [
                    product_data.get('productname'),
                    product_data.get('description'),
                    product_data.get('colors'),
                    product_data.get('size'),
                    product_data.get('sku')
                ])
                result = cursor.fetchone()
                main_product_id = result["id"]
            
           
            if delete_variant_ids:
                delete_query = """
                    DELETE FROM product_list
                    WHERE no = ANY(%s);
                """
                cursor.execute(delete_query, [delete_variant_ids])
            
           
            for variant in generated_variants:
                if "no" in variant:
                    update_variant_query = """
                        UPDATE product_list
                        SET product_name = %s, sku = %s, color = %s, size = %s, imgurl = %s, quantity = %s, price = %s
                        WHERE no = %s;
                    """
                    cursor.execute(update_variant_query, [
                        variant.get('product_name'),
                        variant.get('sku'),
                        variant.get('color'),
                        variant.get('size'),
                        variant.get('imgurl'),
                        variant.get('quantity'),
                        variant.get('price'),
                        variant['no']
                    ])
                else:
                    insert_variant_query = """
                        INSERT INTO product_list (product_name, sku, color, size, imgurl, quantity, price, product_t_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
                    """
                    cursor.execute(insert_variant_query, [
                        variant.get('product_name'),
                        variant.get('sku'),
                        variant.get('color'),
                        variant.get('size'),
                        variant.get('imgurl'),
                        variant.get('quantity'),
                        variant.get('price'),
                        main_product_id
                    ])
            
            conn.commit() 
            return jsonify({"message": "Product and variants created/updated successfully"}), 200
    
    except Exception as e:
        conn.rollback() 
        print(f"Error creating/updating product and variants: {e}")
        return jsonify({"message": "An error occurred while creating/updating product and variants", "error": str(e)}), 500
    
    finally:
        conn.close() 


def delete_product_kind_and_product(id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

       
        delete_products_query = """
            DELETE FROM product_list
            WHERE product_t_id = %s;
        """
        cur.execute(delete_products_query, (id,))

        
        delete_product_kind_query = """
            DELETE FROM productkinds
            WHERE id = %s
            RETURNING *;
        """
        cur.execute(delete_product_kind_query, (id,))
        result = cur.fetchone()

        if result is None:
            return jsonify({"message": "Product Kind not found"}), 404

        conn.commit()
        return jsonify({
            "message": "Product Kind and associated products deleted successfully"
        }), 200

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"message": str(e) or "Some error occurred!"}), 500

    finally:
        if conn:
            cur.close()
            conn.close()


def export_product_data():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

       
        query = """
            SELECT 
                pk.productname AS ProductType, 
                pk.sku AS ProductKindSKU, 
                pk.colors AS Colors, 
                pk.size AS Sizes, 
                pk.description AS Description, 
                p.product_name AS ProductName, 
                p.sku AS SKU, 
                p.color AS Color, 
                p.size AS Size, 
                p.quantity AS Quantity, 
                p.price AS Price
            FROM productkinds pk
            LEFT JOIN product_list p ON pk.id = p.product_t_id
            ORDER BY pk.id ASC;
        """
        cur.execute(query)
        rows = cur.fetchall()

       
        headers = [
            "ProductName", "SKU", "Color", "Size", "Quantity", "Price",
            "ProductType", "ProductKindSKU", "Colors", "Sizes", "Description"
        ]

        combined_data = []
        for row in rows:
            combined_data.append({
                "ProductName": row[5],
                "SKU": row[6],
                "Color": row[7] or "",
                "Size": row[8] or "",
                "Quantity": row[9] or 0,
                "Price": locale.format_string("%.2f", row[10] or 0.0, grouping=True),
                "ProductType": row[0] or "",
                "ProductKindSKU": row[1] or "",
                "Colors": ", ".join(row[2]) if row[2] else "",
                "Sizes": ", ".join(row[3]) if row[3] else "",
                "Description": row[4] or "",
            })

       
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ProductData"

       
        sheet.append(headers)

        
        for data in combined_data:
            sheet.append(list(data.values())) 

       
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

       
        return send_file(
            buffer,
            as_attachment=True,
            download_name="productData.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print("Error exporting product data:", e)
        return jsonify({"message": "Error exporting product data", "error": str(e)}), 500
    finally:
        if conn:
            cur.close()
            conn.close()



def import_product_data():
    try:
       
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded."}), 400
        
        file = request.files['file']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        
        if sheet.max_row <= 1:
            return jsonify({"status": "error", "message": "The Excel file contains no data."}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        skipped_rows = []

        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                ProductName,
                SKU,
                Color,
                Size,
                Quantity,
                Price,
                ProductType,
                ProductKindSKU,
                Colors,
                Sizes,
                Description
            ) = row

            
           

            
            cur.execute(
                """
                SELECT id FROM productkinds WHERE sku = %s
                """, (ProductKindSKU,)
            )
            product_kind = cur.fetchone()

            if not product_kind:
                cur.execute(
                    """
                    INSERT INTO productkinds (productname, sku, colors, size, description)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        ProductType,
                        ProductKindSKU,
                        Colors.split(",") if Colors else None,
                        Sizes.split(",") if Sizes else None,
                        Description
                    )
                )
                product_kind_id = cur.fetchone()[0]
            else:
                product_kind_id = product_kind[0]

            
            cur.execute(
                """
                INSERT INTO product_list (product_name, sku, color, size, quantity, price, product_t_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (ProductName, SKU, Color, Size, Quantity, Price, product_kind_id)
            )

        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Product data imported successfully.",
            "skippedRows": skipped_rows
        }), 200

    except Exception as e:
        print("Error importing product data:", e)
        return jsonify({"status": "error", "message": "Error importing product data.", "error": str(e)}), 500

    finally:
        if conn:
            cur.close()
            conn.close()
